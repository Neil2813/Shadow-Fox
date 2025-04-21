from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# --- Scrape ShadowFox Website ---
url = "https://www.shadowfox.in/index.html"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"
}

response = requests.get(url, headers=headers)
html_content = response.text if response.status_code == 200 else ""

soup = BeautifulSoup(html_content, 'lxml')

# Extract data from page
shadowFox_title = soup.find("div", class_="text-box").text.strip()
shadowFox_Description = soup.find("div", class_="col-md-6 mx-auto").text.strip()
shadowFox_Perks = soup.find("div", class_="container text-center").text.strip()
shadowFox_Internship = "Internships we offer"  # Custom label
shadowFox_Footer = soup.find("div", class_="footer").text.strip()
shadowFox_MadeWith = soup.find("div", class_="made-with").text.strip()
shadowFox_Copyright = soup.find("div", class_="copyright").text.strip()

# Internship data with relative image paths
internships = [
    ("Images/data science.png", "Explore our Data Science program, delving into data analysis, machine learning, and predictive modeling."),
    ("Images/Python.png", "Enroll in our Python Development program to master efficiency and build scalable applications."),
    ("Images/aiml.png", "Join our AI/ML program and unlock the potential of Artificial Intelligence and Machine Learning."),
    ("Images/uiux.png", "Design seamless user experiences in our UI/UX Design program."),
    ("Images/webdev1.png", "Develop dynamic and responsive websites in our Web Development program."),
    ("Images/cyber.png", "Explore our Cybersecurity program to learn the essentials of safeguarding digital landscapes."),
    ("Images/android.png", "Learn Android App Development and create robust, user-friendly mobile applications."),
    ("Images/java.png", "Master Java Development and build powerful backend systems and enterprise-level solutions.")
]

# --- Save to Excel ---
try:
    wb = Workbook()
    ws = wb.active

    # Adjust widths
    ws.column_dimensions[get_column_letter(1)].width = 40
    ws.column_dimensions[get_column_letter(3)].width = 60

    # Write header
    ws.append([
        'shadowFox_title', 'shadowFox_Description', 'shadowFox_Internship',
        'shadowFox_Perks', 'shadowFox_Footer', 'shadowFox_MadeWith', 'shadowFox_Copyright'
    ])

    # Write content
    ws.append([
        shadowFox_title, shadowFox_Description, "",  # Internship column to be filled with images
        shadowFox_Perks, shadowFox_Footer, shadowFox_MadeWith, shadowFox_Copyright
    ])

    # Write internships in the 3rd column (C), starting from row 3
    row = 3
    for img_path, desc in internships:
        ws.cell(row=row, column=3, value=desc)

        # Add image beside the text
        img = ExcelImage(img_path)
        img.width, img.height = 120, 120
        ws.add_image(img, f"C{row}")

        row += 6  # Leave space between images

    wb.save("shadowfox.xlsx")
    print("All data saved to 'shadowfox.xlsx' with embedded images successfully!")

except PermissionError:
    print("Permission denied. Please close 'shadowfox.xlsx' if it's open and try again.")
except Exception as e:
    print("Error occurred:", e)
